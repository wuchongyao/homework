#!/usr/bin/env python
#operation for word
#author:wuchongyao
#time:2018-08-16
#-*- coding:utf-8 -*-


import openpyxl
import os


class excel_xlsx():
    def __init__(self, src_name, ws_name, dst_name):
		'''初始化类'''
        self.wb = openpyxl.load_workbook(src_name)
        self.ws_name = ws_name 
        self.dst_name = dst_name


    def create_sheet(self):
		'''创建worksheet'''
        self.wb.active
        ws1 = self.wb[self.ws_name]
        for row in ws1.rows:
            if row[0].coordinate != 'A1' and row[1].value:
                year = row[0].value[:4]
                if year not in self.wb.sheetnames:
                    self.wb.create_sheet(year)


    def filter_sheet(self):
		'''过滤相应的数据'''
        # self.wb.active
        ws1 = self.wb[self.ws_name]
        for ws_name_tmp in self.wb.sheetnames:
            if ws_name_tmp != self.ws_name:
                ws_tmp1 = self.wb[ws_name_tmp]
                i = 0
                sum_price = 0
                for row in ws1:
                    tmp_list = []
                    tmp_list = [row[0].value, row[1].value]
                    if row[0].coordinate == 'A1' and row[1].value :
                        ws_tmp1.append(tmp_list) 
                    else:
                        year = row[0].value[:4]
                        if year == ws_name_tmp:
                            i = i + 1
                            sum_price = sum_price + int(row[1].value)
                            ws_tmp1.append(tmp_list)
                ws_tmp1.append(['平均价格', sum_price / i])

    
    def save_workbook(self):
		'''保存workbook'''
        self.wb.save(self.dst_name)


def main():
    src_name = r"btc.xlsx"
    os.chdir(r"d:/python/practice")
    dst_name = r"btc03.xlsx"
    wb = excel_xlsx(src_name, 'btc', dst_name)
    wb.create_sheet()
    wb.filter_sheet()
    wb.save_workbook()

if __name__ == '__main__':
    main()
