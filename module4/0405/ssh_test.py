#!/usr/bin/env python
# -*- coding:utf-8 -*-
# Date : 2018-11-12
# Author : wuchongyao
# managet server


import paramiko
import subprocess
import os
import sys
import time
from openpyxl import Workbook
from openpyxl import load_workbook


BASE_PATH, _ = os.path.split(os.path.abspath(__file__))
USER_FILE = BASE_PATH + r'\user_info.xlsx'


def is_file_exists(file_path):
    '''判断用户信息文件是否存在，若不存在则新建'''
    if not os.path.exists(file_path):
        wb = Workbook()
        ws = wb.active
        ws['A1'] = "Hostname"
        ws['B1'] = "Username"
        ws['C1'] = "Password"
        ws.title = "user_info"
        wb.save(file_path)


def inq_hostname(hostname):
    '''匹配主机，并查询用户名与密码'''
    wb = load_workbook(USER_FILE, read_only=True)
    ws = wb['user_info']
    for row in ws.rows:
        if row[0].value == hostname:
            username = row[1].value
            password = row[2].value
            return username, password
        else:
            return False


def ssh_connect(ipaddr, user_name, passwd):
    client = paramiko.SSHClient()
    client.set_missing_host_key_policy(paramiko.AutoAddPolicy)
    try:
        client.connect(hostname=ipaddr, username=user_name, password=passwd)
    except Exception as e:
        raise e
    
    client.close()


def file_operation(hostname, user_name, passwd, localfile, remotefile, method):
    ssh = paramiko.Transport((hostname, 22))
    ssh.connect(username=user_name, password=passwd)
    sftp = paramiko.SFTPClient.from_transport(ssh)
    if method == "u":
        try:
            sftp.put(localfile, remotefile)
        except Exception as e:
            raise e
        finally:
            ssh.close()
    elif method == "d":
        try:
            sftp.get(remotefile, localfile)
        except Exception as e:
            raise e
        finally:
            ssh.close()
    else:
        print("This method is not exist")


def linuxshell(ipaddr, user_name, passwd):
    client = paramiko.SSHClient()
    # while True:
    client.set_missing_host_key_policy(paramiko.AutoAddPolicy)
    client.connect(hostname=ipaddr, username=user_name, password=passwd)
    chan = client.invoke_shell()
    print(f"connect {ipaddr} success...")
    while True:
        cmd = input(f'{ipaddr} ~#:')
        if cmd == "\x03":
            break
        chan.send(cmd + '\r')
        time.sleep(0.5)
        rsp = chan.recv(1024).decode('utf-8')
        print(rsp)


def main(ipaddr, user_name, passwd):
    # is_file_exists(USER_FILE)
    # inq_hostname(USER_FILE)
    linuxshell(ipaddr, user_name, passwd)


if __name__ == '__main__':
    main(sys.argv[1], sys.argv[2], sys.argv[3])
